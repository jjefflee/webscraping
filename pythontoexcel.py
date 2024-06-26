import openpyxl as x1 
from openpyxl.styles import Font

wb = x1.Workbook()

ws = wb.active 

ws.title = 'First Sheet'

ws.create_sheet(index=1, title = 'Second Sheet')

ws['A1'] = 'Invoice'
ws['A2'] = 'Tires'
ws['A3'] = 'Brakes'
ws['A4'] = 'Alignment'

header_font = Font(name='Times New Roman', size = 24, bold=True )
ws['A1'].font = header_font




wb.save('PythonToExcel.xlsx')




wb = x1.workbook
ws = wb.active

write_sheet = wb['Second Sheet']

read_wb = xl.load_workbook('ProduceReport.xlsx')
read_ws = read_wb['ProduceReport']

maxC = read_ws.max_column
maxR = read_ws.max_row 

print(maxC)
print(maxR)


write_sheet['A1'] = 'Produce'
write_sheet['B1'] = 'Cost Per Pound'
write_sheet['C1'] = 'Amt Sold'
write_sheet['D1'] = 'Total'

current_row = 2 
colA = 1
colB = 2 
colC =3 
colD = 4

for row in read_ws.iter_rows(min_row=2,max_row=maxR,max_col=maxC):
    print(row)
    name = row[0].value 
    cost = row[1].value 
    amt_sold = float(row[2].value )
    total = float(row[3].value)
    
    write_sheet.call(current_row,colA).value = name 
    write_sheet.call(current_row,colB).value = cost
    write_sheet.call(current_row,colC).value = amt_sold
    write_sheet.call(current_row,colD).value = total
    
    
    current_row += 1
    
summary_row = current_row + 1

write_sheet['B' + str(summary_row)] = 'Total'
write_sheet['B' + str(summary_row)].font = Font(size=16, bold=True)

write_sheet['C' + str(summary_row)] = '=SUM(C2:C' +str(current_row)+')'
write_sheet['D' + str(summary_row)] = '=SUM(D2:D' +str(current_row)+')'

summary_row += 1 

write_sheet['B' + str(summary_row)] = 'Averages'
write_sheet['C' + str(summary_row)] = Font(size=16,bold=True)

write_sheet['C' + str(summary_row)] = '=Average(C2:C' + str(current_row)+ ')'
write_sheet['D' + str(summary_row)] = '=Average(D2:D' + str(current_row)+ ')'

                                                            


write_sheet.column_dimensions['A'].width =15
write_sheet.column_dimensions['B'].width =15
write_sheet.column_dimensions['C'].width =15
write_sheet.column_dimensions['D'].width =15

for cell in write_sheet["C","C"]:
    cell.number_format = '#,##0'
for cell in write_sheet["D","D"]:
    cell.number_format = u'"'
wb.save('PythonToExcel.xlsx')